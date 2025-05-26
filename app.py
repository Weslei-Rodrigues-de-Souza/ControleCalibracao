import sqlite3 
import datetime
import os
import json
import shutil
import subprocess

import smtplib 
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart 
import requests 
import threading 
import pandas as pd 
from io import BytesIO, StringIO
from zoneinfo import ZoneInfo 
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user

from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, abort, g, jsonify, Response, make_response
from werkzeug.utils import secure_filename
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows 
from openpyxl.utils import get_column_letter 

from database import DatabaseManager 
from werkzeug.security import generate_password_hash, check_password_hash
from database import DatabaseManager # Assumindo que DatabaseManager está em database.py
ANEXOS_BASE_DIR_NAME = "anexos_certificados_flask"
ANEXOS_EMPRESAS_DIR_NAME = "anexos_empresas_iso" 
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
NOTIFICACAO_CONFIG_FILE_NAME = "notificacao_config_flask.json" 
NOTIFICACAO_CONFIG_FILE_PATH = os.path.join(BASE_DIR, NOTIFICACAO_CONFIG_FILE_NAME)
DB_NAME_FOR_PATH = "calibracao_equipamentos.db"
DB_FULL_PATH = os.path.join(BASE_DIR, DB_NAME_FOR_PATH)


app = Flask(__name__) 
app.secret_key = os.urandom(24) 
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, ANEXOS_BASE_DIR_NAME)
app.config['UPLOAD_FOLDER_EMPRESAS'] = os.path.join(BASE_DIR, ANEXOS_EMPRESAS_DIR_NAME) 
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx'} # Adicionado mais extensões


COLOR_RULES_FIXED = [
    {'limite_inferior': 90, 'limite_superior': None, 'cor_hex': '#006400', 'nome': 'Verde Escuro', 'tag_style': 'status_verde_escuro'},
    {'limite_inferior': 45, 'limite_superior': 90, 'cor_hex': '#90EE90', 'nome': 'Verde Claro', 'tag_style': 'status_verde_claro'},
    {'limite_inferior': 15, 'limite_superior': 45, 'cor_hex': '#FFFFE0', 'nome': 'Amarelo', 'tag_style': 'status_amarelo'},
    {'limite_inferior': 1, 'limite_superior': 15, 'cor_hex': '#FFA500', 'nome': 'Laranja', 'tag_style': 'status_laranja'},
    {'limite_inferior': None, 'limite_superior': 1, 'cor_hex': '#FF0000', 'nome': 'Vermelho', 'tag_style': 'status_vermelho'}
]
STATUS_INATIVO_COR_HEX = "#B0B0B0"
STATUS_EM_CALIBRACAO_COR_HEX = "#ADD8E6"
STATUS_SEM_DATA_COR_HEX = "#F5F5F5"

REGRAS_VALIDACAO_CRITERIOS = ["B*3 < A", "B*2 < A", "B < A", "A/3 > B", "A/2 > B", "A > B", "Nenhuma"]
CAMPOS_TABELA_NOTIFICACAO = {
    "id": "ID", "nome": "Nome", "tipo_equipamento_nome": "Tipo Equip.", "tag": "TAG",
    "numero_serie": "Nº Série", "proxima_data_calibracao": "Próx. Cal.",
    "dias_vencimento": "Dias Venc.", "status": "Status", "localizacao": "Localização"
}
CRITERIOS_VENCIMENTO_NOTIFICACAO = [
    "Notificar equipamentos com vencimento em até 30 dias",
    "Notificar equipamentos com vencimento em até 45 dias",
    "Notificar equipamentos com vencimento em até 90 dias",
    "Notificar apenas equipamentos vencidos (<= 0 dias)"
]
CRITERIOS_VENCIMENTO_NOTIFICACAO_MANUAL = ["Usar configuração padrão do sistema"] + CRITERIOS_VENCIMENTO_NOTIFICACAO
PERIODICIDADE_NOTIFICACAO = ["Desativado", "Diário", "Semanal", "Quinzenal", "Mensal", "Bimestral", "Trimestral"]
HORARIOS_NOTIFICACAO = [f"{h:02d}:00" for h in range(0, 24)]

db = DatabaseManager(DB_FULL_PATH) 

# --- RESETAR SENHA DO ADMIN PARA 123 SEMPRE QUE INICIAR (remova depois de testar) ---
admin_user = db.get_user_by_username("Admin")
if not admin_user:
    db.create_user("Admin", generate_password_hash("123", method='pbkdf2:sha256'))
    # Se quiser marcar para troca de senha, faça isso em seguida:
    admin_user = db.get_user_by_username("Admin")
    if admin_user:
        db.set_password_change_required(admin_user['id'], True)
    print("Usuário Admin criado com senha 123.")
else:
    db.update_user_password(admin_user['id'], generate_password_hash("123", method='pbkdf2:sha256'))
    db.set_password_change_required(admin_user['id'], True)
    print("Senha do usuário Admin foi resetada para 123.")

# Inicializar o Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # Define o endpoint para a página de login

# Classe de Usuário para Flask-Login
class User(UserMixin):
    def __init__(self, user_id, nome_usuario, ativo=True, requires_password_change=False):
        self.id = user_id
        self.nome_usuario = nome_usuario
        self._ativo = ativo  # Use atributo privado
        self.requires_password_change = requires_password_change

    @property
    def is_active(self):
        return self._ativo

# Função user_loader para Flask-Login
@login_manager.user_loader
def load_user(user_id):
    """Carrega um usuário dado seu ID.""" 
    db_instance = db  # Use o objeto global db
    user_data = db_instance.get_user_by_id(user_id)
    if user_data:
        user_dict = dict(user_data)
        user = User(
            user_dict['id'],
            user_dict['nome_usuario'],
            user_dict.get('ativo', True),
            user_dict.get('requires_password_change', False)
        )
        return user
    return None

class AppUtils:
    def __init__(self, db_manager_instance):
        self.db_manager = db_manager_instance
        self.regras_cores = COLOR_RULES_FIXED

    @staticmethod
    def format_date_for_display(date_str_iso):
        if not date_str_iso:
            return "N/A"
        try:
            if isinstance(date_str_iso, (datetime.date, datetime.datetime)):
                return date_str_iso.strftime('%d/%m/%Y')
            dt_obj = datetime.datetime.strptime(str(date_str_iso), '%Y-%m-%d')
            return dt_obj.strftime('%d/%m/%Y')
        except ValueError:
            try:
                dt_obj = datetime.datetime.strptime(str(date_str_iso), '%Y-%m-%d %H:%M:%S')
                return dt_obj.strftime('%d/%m/%Y')
            except ValueError:
                 return str(date_str_iso)
        except TypeError:
             return "N/A"


    @staticmethod
    def calcular_dias_para_vencimento(data_proxima_str, ativo=1, em_calibracao=0):
        if not ativo: return None, "status_inativo", STATUS_INATIVO_COR_HEX
        if em_calibracao: return None, "status_em_calibracao", STATUS_EM_CALIBRACAO_COR_HEX
        if not data_proxima_str or data_proxima_str == "N/A":
            return None, "status_semdata", STATUS_SEM_DATA_COR_HEX
        try:
            data_proxima = datetime.datetime.strptime(str(data_proxima_str), "%Y-%m-%d").date()
            hoje = datetime.date.today()
            delta = (data_proxima - hoje).days
            for regra in AppUtils.regras_cores:
                lim_inf, lim_sup = regra['limite_inferior'], regra['limite_superior']
                if lim_inf is not None and lim_sup is not None:
                    if lim_inf <= delta < lim_sup: return delta, regra['tag_style'], regra['cor_hex']
                elif lim_inf is not None and lim_sup is None:
                    if delta >= lim_inf: return delta, regra['tag_style'], regra['cor_hex']
                elif lim_inf is None and lim_sup is not None:
                     if delta < lim_sup: return delta, regra['tag_style'], regra['cor_hex']
            return delta, "status_semdata", STATUS_SEM_DATA_COR_HEX
        except ValueError as e:
            print(f"DEBUG calcular_dias: ValueError ao parsear data '{data_proxima_str}': {e}")
            return None, "status_semdata", STATUS_SEM_DATA_COR_HEX

    def check_calibration_due_dates_and_update_status(self):
        # Obter uma nova instância do DB para esta thread/contexto, se necessário
        equipamentos = self.db_manager.fetch_all_equipamentos_completos()
        updated_count = 0
        for eq_data_row in equipamentos:
            eq_data = dict(eq_data_row)
            eq_id = eq_data['id']
            prox_cal_str_equip = eq_data['proxima_data_calibracao']
            current_status_equip = eq_data['status']
            ativo = eq_data['ativo']
            em_calibracao = eq_data['em_calibracao']

            if not ativo or em_calibracao: continue
            dias_venc, _, _ = AppUtils.calcular_dias_para_vencimento(prox_cal_str_equip, ativo, em_calibracao)
            new_status = current_status_equip
            # Lógica para atualizar o status baseada nos dias_venc
            if dias_venc is not None:
                if dias_venc <= 0 and current_status_equip != "Calibração Vencida":
                    new_status = "Calibração Vencida"
                elif dias_venc > 0 and current_status_equip == "Calibração Vencida":
                     # Se estava vencido e agora não está mais (ex: data corrigida ou nova análise), definir para um status padrão ou baseado em lógica adicional
                     # Aqui, assumimos que se não está vencido, e não está 'Em Calibração', o status padrão é 'Ativo' ou manter o último status não-vencido.
                     # Uma abordagem simples é definir como 'Ativo' se não estiver 'Em Calibração' e a data for futura
                     if not em_calibracao:
                          new_status = "Ativo" # Ou outro status padrão apropriado

            if new_status != current_status_equip:
                 self.db_manager.update_equipamento_status(eq_id, new_status)
                 updated_count += 1
        # print(f"Status de {updated_count} equipamentos atualizados.") # Opcional: logar atualizações

utils = AppUtils(db) # <-- Instancie a classe AppUtils aqui, após sua definição

# Adicionar rotas de Login e Logout
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
            flash('Por favor, altere sua senha temporária.', 'warning') # Mensagem amarela/laranja
            if current_user.requires_password_change:
                flash('Por favor, altere sua senha temporária.', 'warning')
                return redirect(url_for('gerenciar_usuarios'))

            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard')) # Redireciona para o dashboard ou página original se já logado e senha não precisa ser trocada


    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        db_instance = db # Obter a instância do banco de dados usando get_db()
        user_data = db_instance.get_user_by_username(username)
        if user_data and check_password_hash(user_data['senha'], password):
            user_dict = dict(user_data)
            user = User(
                user_dict['id'],
                user_dict['nome_usuario'],
                user_dict.get('ativo', True),
                user_dict.get('requires_password_change', False)
            )
            login_user(user)

            if user.requires_password_change:
                flash('Por favor, altere sua senha temporária.', 'warning')
                return redirect(url_for('gerenciar_usuarios')) # Redireciona para a página de gerenciamento
            else:
                next_page = request.args.get('next') # Pega a URL original que o usuário tentou acessar (pode ser usado, mas o redirecionamento abaixo tem prioridade)
                return redirect(next_page or url_for('dashboard')) # Redireciona

        else:
            # Credenciais inválidas
            flash('Usuário ou senha inválidos.', 'danger') # Use 'danger' para cor vermelha (Bootstrap) ou defina suas classes
            return render_template('login.html')

    # Método GET
    return render_template('login.html')

# Rota de Logout
@app.route('/logout')
@login_required # Apenas usuários logados podem fazer logout
def logout():
    logout_user()
    flash('Você foi desconectado.', 'info') # Mensagem de informação
    return redirect(url_for('login')) # Redireciona para a página de login

@app.route('/gerenciar_usuarios')
@login_required # Protegida por login
def gerenciar_usuarios():
    # Apenas administradores ou usuários com permissão devem acessar esta rota
    # Por enquanto, vamos permitir acesso se o usuário for 'Admin' (exemplo simples)
    if current_user.nome_usuario != 'Admin':
        flash('Você não tem permissão para acessar esta página.', 'danger')
        return redirect(url_for('dashboard')) # Ou outra página apropriada

    db_instance = db # Obter a instância do banco de dados
    users = db_instance.get_all_users()
    # Filtrar o usuário atual da lista para não poder excluir a si mesmo pela interface simples
    users_display = [u for u in users if u['id'] != current_user.id]
    return render_template('gerenciar_usuarios.html', users=users_display)

# Rota para Adicionar Novo Usuário (AJAX/Form)
@app.route('/adicionar_usuario', methods=['POST'])
@login_required
def adicionar_usuario():
    if current_user.nome_usuario != 'Admin':
        flash('Você não tem permissão para realizar esta ação.', 'danger')
        return redirect(url_for('gerenciar_usuarios'))

    username = request.form.get('username')
    # Gerar uma senha temporária e marcar para troca obrigatória
    temporary_password = "changeme123" # Senha temporária simples, idealmente mais complexa ou gerada aleatoriamente
    hashed_password = generate_password_hash(temporary_password, method='pbkdf2:sha256')

    db_instance = db
    try:
        success = db_instance.create_user(username, hashed_password)
        if success:
            # Marcar para troca obrigatória de senha
            novo_usuario = db_instance.get_user_by_username(username)
            if novo_usuario:
                db_instance.set_password_change_required(novo_usuario['id'], True)
            flash(f'Usuário "{username}" criado com sucesso. Senha temporária: "{temporary_password}". Requer troca no primeiro login.', 'success')
        else:
            flash(f'Erro ao criar usuário "{username}". Nome de usuário já existe?', 'danger')
    except Exception as e:
        flash(f'Erro inesperado ao criar usuário: {e}', 'danger')

    return redirect(url_for('gerenciar_usuarios'))

@app.route('/alterar_senha', methods=['POST'])
@login_required
def alterar_senha():
    nova_senha = request.form.get('nova_senha')
    confirmar_senha = request.form.get('confirmar_senha')
    if not nova_senha or nova_senha != confirmar_senha:
        flash('As senhas não coincidem.', 'danger')
        return redirect(url_for('gerenciar_usuarios'))
    db.update_user_password(current_user.id, generate_password_hash(nova_senha, method='pbkdf2:sha256'))
    db.set_password_change_required(current_user.id, False)
    flash('Senha alterada com sucesso!', 'success')
    return redirect(url_for('dashboard'))

@app.context_processor
def inject_utilities():
    return {'now': datetime.datetime.now(datetime.timezone.utc), 'format_date': utils.format_date_for_display}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# --- Rotas Flask ---
@app.route('/')
@login_required
def index(): # Renomeado de dashboard para ser a rota raiz, o dashboard real será outra rota
    print(f"Usuário logado: {current_user.nome_usuario}")
    return render_template('dashboard.html', 
                           total_equip=total_equip, 
                           ativos_count=ativos_count, 
                           em_calibracao_count=em_calibracao_count)
@app.route('/dashboard')
@login_required
def dashboard():
    utils.check_calibration_due_dates_and_update_status()
    equipamentos = db.fetch_all_equipamentos_completos()
    total_equip = len(equipamentos) if equipamentos else 0
    ativos_count = sum(1 for eq in equipamentos if eq['ativo']) if equipamentos else 0
    em_calibracao_count = sum(1 for eq in equipamentos if eq['ativo'] and eq['em_calibracao']) if equipamentos else 0
    return render_template('dashboard.html', total_equip=total_equip, 
                           ativos_count=ativos_count, 
                           em_calibracao_count=em_calibracao_count)

@app.route('/equipamentos')
def lista_equipamentos():
    # Esta rota ainda não está protegida com @login_required para manter a funcionalidade atual.
    search_query = request.args.get('search', '')
    if search_query:
        equipamentos_data = db.search_equipamentos(search_query)
    else:
        equipamentos_data = db.fetch_all_equipamentos_completos()
    
    empresas_calibracao_data = db.fetch_empresas_calibracao() # Use esta função
    empresas_unidade = db.fetch_empresas_unidade()
    equipamentos_display = []

    if equipamentos_data:
        for equip_row in equipamentos_data:
            equip = dict(equip_row) 
            dias, tag_style, cor_hex = utils.calcular_dias_para_vencimento(
                equip.get('proxima_data_calibracao'),
                equip.get('ativo'),
                equip.get('em_calibracao')
            )
            equip['dias_vencimento_display'] = str(dias) if dias is not None else "N/A"
            equip['cor_vencimento'] = cor_hex
            
            if not equip.get('ativo'):
                equip['status_display'] = "Inativo"
                equip['cor_status_bg'] = STATUS_INATIVO_COR_HEX
                equip['cor_status_text'] = "white" if STATUS_INATIVO_COR_HEX == "#B0B0B0" else "black"
            elif equip.get('em_calibracao'):
                equip['status_display'] = "Em Calibração"
                equip['cor_status_bg'] = STATUS_EM_CALIBRACAO_COR_HEX
                equip['cor_status_text'] = "black"
            else:
                equip['status_display'] = equip.get('status', 'N/A')
                if equip['status_display'] == "Calibração Vencida":
                     equip['cor_status_text'] = "red"
                else: 
                     equip['cor_status_text'] = "black" 
                     equip['cor_status_bg'] = "transparent" 

            equipamentos_display.append(equip)
    
    tipos_equip_para_modal = db.fetch_all_tipos_equipamento()

    return render_template('lista_equipamentos.html', 
                           equipamentos=equipamentos_display, 
                           search_query=search_query,
                           empresas=empresas_unidade, # Manter para outros usos na página, se necessário
                           empresas_calibracao=[dict(row) for row in empresas_calibracao_data], # Passar a lista de empresas de calibração
                           tipos_equip_para_modal=tipos_equip_para_modal,
                           REGRAS_VALIDACAO_CRITERIOS=REGRAS_VALIDACAO_CRITERIOS) 

@app.route('/equipamento/novo', methods=['POST'])
@login_required
def novo_equipamento():
    if request.method == 'POST':
        dados_equip = {
            'nome': request.form.get('nome'),
            'fabricante': request.form.get('fabricante'),
            'modelo': request.form.get('modelo'),
            'numero_serie': request.form.get('numero_serie'),
            'tag': request.form.get('tag'),
            'status': request.form.get('status_lista'),
            'localizacao': request.form.get('localizacao'),
            'empresa_id': request.form.get('empresa_id') if request.form.get('empresa_id') not in [None, "None", ""] else None,
            'observacoes_equipamento': request.form.get('observacoes_equipamento'),
            'tipo_equipamento_id': request.form.get('tipo_equipamento_id') if request.form.get('tipo_equipamento_id') not in [None, "None", ""] else None,
            'faixa_de_uso': request.form.get('faixa_de_uso'),
            'ativo': 'ativo' in request.form,
            'requer_calibracao': 'requer_calibracao' in request.form,
            'em_calibracao': 'em_calibracao' in request.form,
            'destino_inativo': request.form.get('destino_inativo') if 'ativo' not in request.form else None
        }
        if not dados_equip['ativo']:
            dados_equip['status'] = "Inativo"
        elif dados_equip['em_calibracao']:
            dados_equip['status'] = "Em Calibração"

        if not dados_equip['nome']:
            flash('O nome do equipamento é obrigatório.', 'danger')
        else:
            try:
                if db.add_equipamento(dados_equip):
                    flash('Equipamento adicionado com sucesso!', 'success')
                else:
                    if not any(message for category, message in get_flashed_messages(with_categories=True) if category == 'danger'):
                        flash('Erro ao adicionar equipamento.', 'danger')
            except Exception as e:
                flash(f"Erro ao adicionar equipamento: {e}", "danger")
    return redirect(url_for('lista_equipamentos')) 

@app.route('/equipamento/<int:equip_id>')
@login_required
def editar_equipamento(equip_id): # Mantém o nome da função, mas a rota é GET
    equip_data_row = db.fetch_equipamento_completo_by_id(equip_id)
    if not equip_data_row:
        return jsonify({"error": "Equipamento não encontrado"}), 404

    equip_data = dict(equip_data_row)
    # Garante que 'empresa_id' está presente no dicionário do equipamento, mesmo que seja None
    equip_data['empresa_id'] = equip_data.get('empresa_id') 

    # Buscar a lista completa de empresas
    empresas_unidade_data = db.fetch_empresas_unidade() # Buscar apenas empresas 'Unidade'
    empresas_list = [dict(row) for row in empresas_unidade_data] # Usar a lista de empresas 'Unidade'

    unidades_medida = []
    if equip_data.get('tipo_equipamento_id'):
        unidades_medida = [dict(row) for row in db.fetch_unidades_by_tipo_id(equip_data['tipo_equipamento_id'])]
    equip_data['unidades_medida_tipo'] = unidades_medida
    
    tipos_equipamento = db.fetch_all_tipos_equipamento() 
    return jsonify({
        "equipamento": equip_data,
        "empresas": empresas_list,  # Incluir a lista de empresas
        "tipos_equipamento": [dict(row) for row in tipos_equipamento]
    })
@app.route('/analise/json/<int:analise_id>')
@login_required
def get_analise_json(analise_id):
    analise_data = db.fetch_analise_by_id(analise_id, app_utils_instance=utils) 
    if not analise_data: 
        return jsonify({"error": "Análise não encontrada"}), 404
    
    analise_data_dict = dict(analise_data) 
    analise_data_dict['anexos'] = [dict(anexo_row) for anexo_row in db.fetch_anexos_by_analise_id(analise_id)]
    
    equip_associado = db.fetch_equipamento_completo_by_id(analise_data_dict['equipamento_id'])
    unidades_medida_analise = []
    if equip_associado and equip_associado['tipo_equipamento_id']:
        unidades_medida_analise = [dict(row) for row in db.fetch_unidades_by_tipo_id(equip_associado['tipo_equipamento_id'])]
    analise_data_dict['unidades_medida_tipo'] = unidades_medida_analise
    
    analise_data_dict['pontos_analisados'] = [dict(p_row) for p_row in db.fetch_pontos_by_analise_id(analise_id)]

    return jsonify({"analise": analise_data_dict})

@app.route('/tipo/json/<int:tipo_id>') 
@login_required
def tipo_json(tipo_id):
    tipo_data = dict(tipo_data_row)
    empresas_data = db.fetch_all_empresas() 
    tipo_data['empresas'] = [dict(row) for row in empresas_data] 

    unidades_data = [dict(row) for row in db.fetch_unidades_by_tipo_id(tipo_id)]
    return jsonify({"tipo": tipo_data, "unidades": unidades_data})

@app.route('/equipamento/editar/<int:equip_id>', methods=['POST']) 
@login_required
def post_editar_equipamento(equip_id):
    try:
        equip_data_row = db.fetch_equipamento_completo_by_id(equip_id)
        if not equip_data_row:
            return jsonify({"success": False, "message": "Equipamento não encontrado."}), 404

        if request.method == 'POST':
            dados_atualizados = {
                'nome': request.form.get('edit_nome'),
                'fabricante': request.form.get('edit_fabricante'),
                'modelo': request.form.get('edit_modelo'),
                'numero_serie': request.form.get('edit_numero_serie'),
                'tag': request.form.get('edit_tag'),
                'status': request.form.get('edit_status_lista'),
                'localizacao': request.form.get('edit_localizacao'),
                'empresa_id': request.form.get('edit_empresa_id') if request.form.get('edit_empresa_id') not in [None, "None", ""] else None,
                'observacoes_equipamento': request.form.get('edit_observacoes_equipamento'),
                'tipo_equipamento_id': request.form.get('edit_tipo_equipamento_id') if request.form.get('edit_tipo_equipamento_id') not in [None, "None", ""] else None,
                'faixa_de_uso': request.form.get('edit_faixa_de_uso'),
                'ativo': 'edit_ativo' in request.form,
                'requer_calibracao': 'edit_requer_calibracao' in request.form,
                'em_calibracao': 'edit_em_calibracao' in request.form,
                'destino_inativo': request.form.get('edit_destino_inativo') if 'edit_ativo' not in request.form else None
            }
            if not dados_atualizados['ativo']:
                dados_atualizados['status'] = "Inativo"
            elif dados_atualizados['em_calibracao']:
                dados_atualizados['status'] = "Em Calibração"

        if not dados_atualizados['nome']:
            return jsonify(success=False, message="O nome do equipamento é obrigatório."), 400
        if db.update_equipamento_principal(equip_id, dados_atualizados):
            return jsonify(success=True, message="Equipamento atualizado com sucesso!")
        else:
            return jsonify(success=False, message="Erro ao atualizar equipamento."), 400
    except Exception as e:
        return jsonify(success=False, message=f"Erro ao atualizar equipamento: {e}"), "danger"

@app.route('/equipamento/excluir/<int:equip_id>', methods=['POST'])
@login_required
def excluir_equipamento(equip_id):
    if db.delete_equipamento(equip_id, app.config['UPLOAD_FOLDER'], app_utils_instance=utils):
        flash('Equipamento excluído com sucesso!', 'success')
    else:
        flash('Erro ao excluir equipamento.', 'danger')
    return redirect(url_for('lista_equipamentos'))

@app.route('/tipos', methods=['GET'])
@login_required
def gerenciar_tipos():
    tipos_raw = db.fetch_all_tipos_equipamento()
    tipos_com_unidades = []
    if tipos_raw:
        for tipo_row in tipos_raw:
            tipo = dict(tipo_row)
            tipo['unidades'] = [dict(u_row) for u_row in db.fetch_unidades_by_tipo_id(tipo['id'])]
            tipos_com_unidades.append(tipo)
    return render_template('gerenciar_tipos.html', tipos=tipos_com_unidades)


@app.route('/tipo/salvar', methods=['POST'])
@app.route('/tipo/salvar/<int:tipo_id>', methods=['POST'])
@login_required
def salvar_tipo(tipo_id=None):
    nome_tipo = request.form.get('nome_tipo')
    unidades_json_data = request.form.get('unidades_json_data', '[]') 
    
    if not nome_tipo:
        flash("O nome do tipo é obrigatório.", "danger")
        return redirect(url_for('gerenciar_tipos'))

    try:
        unidades_para_salvar = json.loads(unidades_json_data)
    except json.JSONDecodeError:
        flash("Erro ao processar dados das unidades.", "danger")
        return redirect(url_for('gerenciar_tipos'))

    if tipo_id is None: 
        try:
            novo_tipo_id = db.add_tipo_equipamento(nome_tipo)
            if novo_tipo_id:
                for unidade in unidades_para_salvar:
                    if unidade.get('status') != 'deleted' and unidade.get('nome_unidade'): 
                        db.add_unidade_medida_config(novo_tipo_id, unidade['nome_unidade'], unidade.get('simbolo_unidade'))
                flash(f"Tipo '{nome_tipo}' adicionado com sucesso!", "success")
            else:
                flash(f"Erro ao adicionar tipo '{nome_tipo}'. Verifique se já existe.", "danger")
        except Exception as e:
            flash(f"Erro ao salvar novo tipo: {e}", "danger")
    else: 
        try:
            if db.update_tipo_equipamento(tipo_id, nome_tipo):
                for unidade_form in unidades_para_salvar:
                    form_unit_id = unidade_form.get('id') 
                    form_unit_status = unidade_form.get('status')
                    
                    if form_unit_status == 'deleted' and form_unit_id is not None:
                        db.delete_unidade_medida_config(form_unit_id)
                    elif form_unit_status == 'new' and unidade_form.get('nome_unidade'):
                        db.add_unidade_medida_config(tipo_id, unidade_form['nome_unidade'], unidade_form.get('simbolo_unidade'))
                flash(f"Tipo '{nome_tipo}' atualizado com sucesso!", "success")
            else:
                flash(f"Erro ao atualizar tipo '{nome_tipo}'. Verifique se o novo nome já existe.", "danger")
        except Exception as e:
            flash(f"Erro ao salvar alterações do tipo: {e}", "danger")
            
    return redirect(url_for('gerenciar_tipos'))

@app.route('/tipos/adicionar_ajax', methods=['POST'])
@login_required
def salvar_tipo_ajax():
    data = request.get_json()
    nome_tipo = data.get('nome_tipo')

    if not nome_tipo:
        return jsonify({"success": False, "message": "O nome do tipo é obrigatório."}), 400
    
    try:
        novo_tipo_id = db.add_tipo_equipamento(nome_tipo)
        if novo_tipo_id:
            tipo_criado = db.fetch_tipo_equipamento_by_id(novo_tipo_id) 
            return jsonify({"success": True, "message": "Tipo adicionado com sucesso!", "tipo": dict(tipo_criado)})
        else: 
            return jsonify({"success": False, "message": "Erro ao adicionar tipo. Verifique se já existe."}), 400
    except Exception as e:
        return jsonify({"success": False, "message": f"Erro ao salvar novo tipo: {str(e)}"}), 500


@app.route('/tipo/excluir/<int:tipo_id>', methods=['POST'])
@login_required
def excluir_tipo(tipo_id):
    unidades_associadas = db.fetch_unidades_by_tipo_id(tipo_id)
    for unidade in unidades_associadas:
        db.delete_unidade_medida_config(unidade['id'])
        
    resultado = db.delete_tipo_equipamento(tipo_id) 
    if resultado == "EM_USO": 
        flash("Não é possível excluir o tipo. Ele está associado a um ou mais equipamentos.", "danger")
    elif resultado: 
        flash("Tipo de equipamento e suas unidades associadas foram excluídos com sucesso.", "success")
    else:
        flash("Erro ao excluir tipo de equipamento.", "danger") 
    return redirect(url_for('gerenciar_tipos'))

# --- Rotas para Empresas ---
@app.route('/empresas')
@login_required
def gerenciar_empresas():
    empresas_data = db.fetch_all_empresas()
    return render_template('gerenciar_empresas.html', empresas=empresas_data)

@app.route('/empresa/salvar', methods=['POST'])
@app.route('/empresa/salvar/<int:empresa_id>', methods=['POST'])
@login_required
def salvar_empresa(empresa_id=None):
    if request.method == 'POST':
        dados_empresa = {
            'razao_social': request.form.get('empresa_razao_social'),
            'nome_fantasia': request.form.get('empresa_nome_fantasia'),
            'cnpj': request.form.get('empresa_cnpj', '').replace('.', '').replace('/', '').replace('-', ''),
            'logradouro': request.form.get('empresa_logradouro'),
            'numero': request.form.get('empresa_numero'),
            'complemento': request.form.get('empresa_complemento'),
            'bairro': request.form.get('empresa_bairro'),
            'cep': request.form.get('empresa_cep', '').replace('-', ''),
            'municipio': request.form.get('empresa_municipio'),
            'uf': request.form.get('empresa_uf'),
            'telefone': request.form.get('empresa_telefone'),
            'email': request.form.get('empresa_email'),
            'categoria': request.form.get('empresa_categoria'),
            'certificado_iso_path': None 
        }
        remover_certificado_existente = request.form.get('remover_certificado_iso_atual') == '1'

        if not dados_empresa['cnpj'] or not dados_empresa['categoria']:
            flash('CNPJ e Categoria são obrigatórios.', 'danger')
            return redirect(url_for('gerenciar_empresas'))

        certificado_file = request.files.get('empresa_certificado_iso')
        
        empresa_atual = None
        if empresa_id:
            empresa_atual = db.fetch_empresa_by_id(empresa_id)
            if empresa_atual:
                dados_empresa['certificado_iso_path'] = empresa_atual['certificado_iso_path']

        # Remover certificado existente se marcado
        if remover_certificado_existente and empresa_atual and empresa_atual['certificado_iso_path']:
            caminho_antigo = os.path.join(app.config['UPLOAD_FOLDER_EMPRESAS'], empresa_atual['certificado_iso_path'])
            if os.path.exists(caminho_antigo):
                try:
                    os.remove(caminho_antigo)
                    dir_empresa_antigo = os.path.dirname(caminho_antigo)
                    if os.path.exists(dir_empresa_antigo) and not os.listdir(dir_empresa_antigo):
                        os.rmdir(dir_empresa_antigo)
                except OSError as e:
                    print(f"Aviso: Erro ao excluir arquivo/pasta do certificado antigo da empresa {empresa_id}: {e}")
            dados_empresa['certificado_iso_path'] = None # Limpa o path no DB

        # Salvar novo certificado se enviado
        if certificado_file and certificado_file.filename != '':
            if allowed_file(certificado_file.filename):
                # Se já existe um certificado e um novo está sendo enviado, remove o antigo primeiro
                if dados_empresa['certificado_iso_path'] and not remover_certificado_existente: # Não removeu explicitamente, mas está substituindo
                    caminho_antigo_para_substituir = os.path.join(app.config['UPLOAD_FOLDER_EMPRESAS'], dados_empresa['certificado_iso_path'])
                    if os.path.exists(caminho_antigo_para_substituir):
                        try:
                            os.remove(caminho_antigo_para_substituir)
                            dir_empresa_subs = os.path.dirname(caminho_antigo_para_substituir)
                            if os.path.exists(dir_empresa_subs) and not os.listdir(dir_empresa_subs):
                                os.rmdir(dir_empresa_subs)
                        except OSError as e:
                             print(f"Aviso: Erro ao excluir arquivo/pasta do certificado antigo da empresa {empresa_id} ao substituir: {e}")
                
                filename_original = secure_filename(certificado_file.filename)
                ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S%f")
                filename_armazenado = f"{ts}_{filename_original}"
                
                # O ID da empresa é necessário para o caminho
                id_para_pasta = empresa_id
                if id_para_pasta is None: # Se for uma nova empresa, precisa salvar primeiro para obter o ID
                    # Salva sem o path do certificado primeiro
                    temp_dados = dados_empresa.copy()
                    temp_dados.pop('certificado_iso_path', None) 
                    id_para_pasta = db.add_empresa(temp_dados)
                    if not id_para_pasta: # Falha ao adicionar a empresa
                         flash('Erro crítico ao tentar adicionar empresa antes de salvar anexo.', 'danger')
                         return redirect(url_for('gerenciar_empresas'))
                
                caminho_relativo_dir_empresa = str(id_para_pasta)
                caminho_relativo_completo = os.path.join(caminho_relativo_dir_empresa, filename_armazenado)
                dir_empresa_absoluto = os.path.join(app.config['UPLOAD_FOLDER_EMPRESAS'], caminho_relativo_dir_empresa)
                os.makedirs(dir_empresa_absoluto, exist_ok=True)
                caminho_destino_absoluto = os.path.join(dir_empresa_absoluto, filename_armazenado)
                certificado_file.save(caminho_destino_absoluto)
                dados_empresa['certificado_iso_path'] = caminho_relativo_completo
            else:
                flash('Tipo de arquivo de certificado ISO não permitido.', 'danger')
                return redirect(url_for('gerenciar_empresas'))
        
        try:
            if empresa_id is None: 
                # Se o ID foi obtido acima (id_para_pasta), então é uma atualização do path do certificado
                if 'id_para_pasta' in locals() and id_para_pasta:
                    db.update_empresa(id_para_pasta, dados_empresa) # Atualiza com o path
                else: # Caso contrário, é uma nova inserção completa (sem anexo inicial)
                    db.add_empresa(dados_empresa)
                flash('Empresa adicionada com sucesso!', 'success')
            else: 
                db.update_empresa(empresa_id, dados_empresa)
                flash('Empresa atualizada com sucesso!', 'success')
        except sqlite3.IntegrityError as e: 
             flash(f"Erro de integridade ao salvar empresa: CNPJ {dados_empresa['cnpj']} já existe. ({e})", "danger")
        except Exception as e:
            flash(f"Erro ao salvar empresa: {e}", "danger")
        
        return redirect(url_for('gerenciar_empresas'))
    return redirect(url_for('gerenciar_empresas'))


@app.route('/empresa/json/<int:empresa_id>')
@login_required
def empresa_json(empresa_id):
    empresa_data = db.fetch_empresa_by_id(empresa_id)
    if empresa_data:
        return jsonify(dict(empresa_data))
    return jsonify({"error": "Empresa não encontrada"}), 404

@app.route('/empresa/excluir/<int:empresa_id>', methods=['POST'])
@login_required
def excluir_empresa(empresa_id):
    if db.delete_empresa(empresa_id, app.config['UPLOAD_FOLDER_EMPRESAS']):
        flash('Empresa excluída com sucesso!', 'success')
    else:
        flash('Erro ao excluir empresa.', 'danger')
    return redirect(url_for('gerenciar_empresas'))

@app.route('/consultar_cnpj/<cnpj>')
@login_required
def consultar_cnpj(cnpj):
    cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
    if len(cnpj_limpo) != 14:
        return jsonify({"error": "CNPJ inválido. Deve conter 14 dígitos."}), 400

    try:
        response = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_limpo}", timeout=10)
        response.raise_for_status()
        data = response.json()
        empresa_info = {
            "razao_social": data.get("razao_social"),
            "nome_fantasia": data.get("nome_fantasia"),
            "logradouro": data.get("logradouro"),
            "numero": data.get("numero"),
            "complemento": data.get("complemento"),
            "bairro": data.get("bairro"),
            "cep": data.get("cep", "").replace(".", "").replace("-", ""),
            "municipio": data.get("municipio"),
            "uf": data.get("uf"),
            "telefone": data.get("ddd_telefone_1") or data.get("ddd_telefone_2"),
            "email": data.get("email")
        }
        return jsonify(empresa_info)
    except requests.exceptions.HTTPError as http_err:
        if http_err.response.status_code == 404:
            return jsonify({"error": "CNPJ não encontrado na BrasilAPI."}), 404
        return jsonify({"error": f"Erro HTTP ao consultar CNPJ: {http_err}"}), 500
    except requests.exceptions.RequestException as e:
        return jsonify({"error": f"Erro ao consultar CNPJ: {e}"}), 500

@app.route('/anexos_empresas/<path:subpath>')
@login_required
def servir_anexo_empresa(subpath):
    return send_from_directory(app.config['UPLOAD_FOLDER_EMPRESAS'], subpath)

@app.route('/anexos/<path:subpath>')
@login_required
def servir_anexo(subpath):
    return send_from_directory(app.config['UPLOAD_FOLDER'], subpath)

@app.route('/configuracoes', methods=['GET', 'POST'])
# @login_required # Manter sem login_required por enquanto para teste das configs
def configuracoes_notificacao():
    # Esta rota ainda não está protegida com @login_required para manter a funcionalidade atual.
    app_utils_instance = AppUtils(db)
    settings = app_utils_instance.load_notification_settings()

    if request.method == 'POST':
        # Campos de E-mail
        remetente_email = request.form.get('remetente_email', settings['remetente_email'])
        remetente_senha = request.form.get('remetente_senha', settings['remetente_senha'])
        para = request.form.get('para', settings['para'])
        cc = request.form.get('cc', settings['cc'])
        assunto = request.form.get('assunto', settings['assunto'])
        corpo_template_email = request.form.get('corpo_template_email', settings['corpo_template_email'])
        
        # Campos de WhatsApp/Gemini
        zapi_instancia = request.form.get('zapi_instancia', settings['zapi_instancia'])
        zapi_token_instancia = request.form.get('zapi_token_instancia', settings['zapi_token_instancia'])
        zapi_client_token = request.form.get('zapi_client_token', settings['zapi_client_token'])
        gemini_api_key = request.form.get('gemini_api_key', settings['gemini_api_key'])
        whatsapp_para = request.form.get('whatsapp_para', settings['whatsapp_para'])
        corpo_template_whatsapp = request.form.get('corpo_template_whatsapp', settings['corpo_template_whatsapp'])

        # Critérios e Agendamento
        criterio_padrao_vencimento = request.form.get('criterio_padrao_vencimento', settings['criterio_padrao_vencimento'])
        agendamento_periodicidade = request.form.get('agendamento_periodicidade', settings['agendamento_periodicidade'])
        agendamento_data_inicio = request.form.get('agendamento_data_inicio', settings['agendamento_data_inicio'])
        agendamento_horario = request.form.get('agendamento_horario', settings['agendamento_horario'])
        criterio_email_manual = request.form.get('criterio_email_manual', settings['criterio_email_manual'])
        criterio_wpp_manual = request.form.get('criterio_wpp_manual', settings['criterio_wpp_manual'])

        # Campos da Tabela para Notificação
        campos_tabela_selecionados = {}
        for key in CAMPOS_TABELA_NOTIFICACAO.keys():
            # Verifica se o campo está presente no form (significa que foi marcado)
            campos_tabela_selecionados[key] = request.form.get(f'campo_tabela_{key}', 'off') == 'on'
            
        settings = {
            "remetente_email": remetente_email,
            "remetente_senha": remetente_senha,
            "para": para,
            "cc": cc,
            "assunto": assunto,
            "corpo_template_email": corpo_template_email,
            "zapi_instancia": zapi_instancia,
            "zapi_token_instancia": zapi_token_instancia,
            "zapi_client_token": zapi_client_token,
            "gemini_api_key": gemini_api_key,
            "whatsapp_para": whatsapp_para,
            "corpo_template_whatsapp": corpo_template_whatsapp,
            "criterio_padrao_vencimento": criterio_padrao_vencimento,
            "agendamento_periodicidade": agendamento_periodicidade,
            "agendamento_data_inicio": agendamento_data_inicio,
            "agendamento_horario": agendamento_horario,
            "criterio_email_manual": criterio_email_manual,
            "criterio_wpp_manual": criterio_wpp_manual,
            "campos_tabela": campos_tabela_selecionados # Atualiza o dicionário de campos selecionados
        }
        settings['zapi_instancia'] = request.form.get('zapi_instancia', settings['zapi_instancia'])
        settings['zapi_token_instancia'] = request.form.get('zapi_token_instancia', settings['zapi_token_instancia'])
        settings['zapi_client_token'] = request.form.get('zapi_client_token', settings['zapi_client_token'])
        settings['gemini_api_key'] = request.form.get('gemini_api_key', settings['gemini_api_key'])
        settings['whatsapp_para'] = request.form.get('whatsapp_para', settings['whatsapp_para'])
        settings['corpo_template_whatsapp'] = request.form.get('corpo_template_whatsapp', settings['corpo_template_whatsapp'])

        settings['criterio_padrao_vencimento'] = request.form.get('criterio_padrao_vencimento', settings['criterio_padrao_vencimento'])
        settings['agendamento_periodicidade'] = request.form.get('agendamento_periodicidade', settings['agendamento_periodicidade'])
        settings['agendamento_data_inicio'] = request.form.get('agendamento_data_inicio', settings['agendamento_data_inicio'])
        settings['agendamento_horario'] = request.form.get('agendamento_horario', settings['agendamento_horario'])
        settings['criterio_email_manual'] = request.form.get('criterio_email_manual', settings['criterio_email_manual'])
        settings['campos_tabela'] = campos_tabela_selecionados

        try:
            with open(NOTIFICACAO_CONFIG_FILE_PATH, 'w') as f:
                json.dump(settings, f, indent=4)
            flash("Configurações de notificação salvas!", "success")
        except IOError as e:
            flash(f"Não foi possível salvar as configurações: {e}", "danger")
        return redirect(url_for('configuracoes_notificacao')) 

    return render_template('config_notificacoes.html', settings=settings, 
                           campos_tabela_notificacao=CAMPOS_TABELA_NOTIFICACAO,
                           criterios_vencimento=CRITERIOS_VENCIMENTO_NOTIFICACAO,
                           criterios_vencimento_manual=CRITERIOS_VENCIMENTO_NOTIFICACAO_MANUAL,
                           periodicidades=PERIODICIDADE_NOTIFICACAO,
                           horarios=HORARIOS_NOTIFICACAO
                           )

# --- Rota para Envio de E-mail Manual ---
@app.route('/enviar_notificacao_email_manual', methods=['POST'])
@login_required
def enviar_notificacao_email_manual():
    settings = utils.load_notification_settings()
    criterio_selecionado = request.form.get('criterio_email_manual', settings['criterio_email_manual']) 

    if not all([settings.get('remetente_email'), settings.get('remetente_senha'), settings.get('para')]):
        return jsonify({"success": False, "message": "Configurações de e-mail (remetente, senha, destinatário) incompletas."}), 400

    dias_limite = None
    apenas_vencidos = False
    if "30 dias" in criterio_selecionado: dias_limite = 30
    elif "45 dias" in criterio_selecionado: dias_limite = 45
    elif "90 dias" in criterio_selecionado: dias_limite = 90
    elif "apenas equipamentos vencidos" in criterio_selecionado: apenas_vencidos = True
    
    equipamentos_todos = db.fetch_all_equipamentos_completos()
    equipamentos_para_notificar = []

    if equipamentos_todos:
        for equip_row in equipamentos_todos:
            equip = dict(equip_row)
            if not equip.get('ativo') or equip.get('em_calibracao'):
                continue
            
            dias_venc, _, _ = utils.calcular_dias_para_vencimento(
                equip.get('proxima_data_calibracao'),
                equip.get('ativo'),
                equip.get('em_calibracao')
            )
            if dias_venc is not None:
                if apenas_vencidos and dias_venc <= 0:
                    equipamentos_para_notificar.append(equip)
                elif dias_limite is not None and 0 < dias_venc <= dias_limite: 
                    equipamentos_para_notificar.append(equip)
                elif dias_limite is not None and dias_venc <= 0 : 
                    equipamentos_para_notificar.append(equip)


    if not equipamentos_para_notificar:
        return jsonify({"success": True, "message": "Nenhum equipamento encontrado para notificação com o critério selecionado."})

    tabela_html = "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'><thead><tr>"
    colunas_selecionadas = [key for key, val in settings.get('campos_tabela', {}).items() if val]
    for col_key in colunas_selecionadas:
        tabela_html += f"<th>{CAMPOS_TABELA_NOTIFICACAO.get(col_key, col_key)}</th>"
    tabela_html += "</tr></thead><tbody>"

    for equip in equipamentos_para_notificar:
        tabela_html += "<tr>"
        for col_key in colunas_selecionadas:
            valor = equip.get(col_key, 'N/D')
            if col_key == "proxima_data_calibracao":
                valor = utils.format_date_for_display(valor)
            elif col_key == "dias_vencimento": 
                 dias_v, _, _ = utils.calcular_dias_para_vencimento(equip.get('proxima_data_calibracao'), equip.get('ativo'), equip.get('em_calibracao'))
                 valor = dias_v if dias_v is not None else "N/A"
            tabela_html += f"<td>{valor}</td>"
        tabela_html += "</tr>"
    tabela_html += "</tbody></table>"

    corpo_email = settings.get('corpo_template_email', "").replace("{tabela_equipamentos}", tabela_html)

    msg = MIMEMultipart('alternative')
    msg['Subject'] = settings.get('assunto', "Alerta de Calibrações")
    msg['From'] = settings['remetente_email']
    msg['To'] = settings['para']
    if settings.get('cc'):
        msg['Cc'] = settings['cc']
    
    msg.attach(MIMEText(corpo_email, 'html', 'utf-8'))

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server: 
            server.login(settings['remetente_email'], settings['remetente_senha'])
            server.send_message(msg)
        return jsonify({"success": True, "message": f"E-mail de notificação enviado para {settings['para']}!"})
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")
        return jsonify({"success": False, "message": f"Erro ao enviar e-mail: {str(e)}"}), 500

# --- Rota para Envio de WhatsApp Manual ---
def _gerar_tabela_texto_para_whatsapp(equipamentos_lista, campos_selecionados_config):
    texto_final = ""
    for equip in equipamentos_lista:
        tipo_equip = equip.get('tipo_equipamento_nome', 'Equipamento')
        texto_final += f"*{tipo_equip.upper()}*\n" 

        if campos_selecionados_config.get("nome"):
            texto_final += f"  * Nome: {equip.get('nome', 'N/D')}\n"
        if campos_selecionados_config.get("tag"): 
            texto_final += f"  * TAG: {equip.get('tag', 'N/D')}\n"
        if campos_selecionados_config.get("numero_serie"):
            texto_final += f"  * Nº Série: {equip.get('numero_serie', '(Não informado)')}\n"
        if campos_selecionados_config.get("proxima_data_calibracao"):
            texto_final += f"  * Próxima Calibração: {utils.format_date_for_display(equip.get('proxima_data_calibracao'))}\n"
        if campos_selecionados_config.get("dias_vencimento"):
            dias_v, _, _ = utils.calcular_dias_para_vencimento(equip.get('proxima_data_calibracao'), equip.get('ativo'), equip.get('em_calibracao'))
            dias_texto = "N/A"
            if dias_v is not None:
                if dias_v < 0:
                    dias_texto = f"{abs(dias_v)} dia(s) vencido(s)"
                elif dias_v == 0:
                    dias_texto = "Vence Hoje"
                else:
                    dias_texto = f"Vence em {dias_v} dia(s)"
            texto_final += f"  * Dias para Venc.: {dias_texto}\n"
        if campos_selecionados_config.get("status"):
            texto_final += f"  * Status: {equip.get('status', 'N/D')}\n"
        if campos_selecionados_config.get("localizacao"):
            texto_final += f"  * Localização: {equip.get('localizacao', 'N/D')}\n"
        texto_final += "\n" 
    return texto_final.strip()

def _gerar_mensagem_whatsapp_com_gemini(tabela_texto, api_key, settings):
    prompt = (
        "Você é um assistente responsável por notificar sobre calibrações de equipamentos.\n"
        "Gere uma mensagem de WhatsApp amigável e profissional informando sobre os equipamentos abaixo que precisam de atenção. Não precisa ser tão formal\n"
        "Use formatação do WhatsApp como *negrito* para destacar informações importantes e _itálico_ se apropriado.\n"
        "Inclua uma saudação cordial e uma despedida.\n\n"
        "Equipamentos para Notificação:\n"
        f"{tabela_texto}\n\n"
        "Use o seguinte modelo para os dados de cada equipamento:\n"
        "* [TIPO DO EQUIPAMENTO (MAIÚSCULAS)]\n"
        "[Lista dos campos selecionados com seus valores no molde Campo: Valor]\n"
        "\n"
        "Lembre-se de manter a mensagem concisa e clara para o WhatsApp.\n"
        "Não indique no texto o nome e departamento e tão pouco que se trata de um e-mail. Coloque sempre ao final: Mensagem automática\n"
        "Padronize a estrutura dos equipamentos no seguinte molde: exemplo:\n"
        "* MULTÍMETRO\n"
        "    * Nome: PAQ-01\n"
        "    * TAG: PAT-001\n" 
        "    * Nº Série: (Não informado)\n"
        "    * Próxima Calibração: 20/05/2025\n"
        "    * Dias Vencidos: 1 dia\n" 
        "    * Status: Calibração Vencida\n"
        "    * Localização: Laboratório"
    )
    
    gemini_api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent?key={api_key}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}]
    } 
    headers = {'Content-Type': 'application/json'}
    
    try:
        response = requests.post(gemini_api_url, headers=headers, json=payload, timeout=45) 
        response.raise_for_status() 
        result = response.json()
        
        if result.get("candidates") and result["candidates"][0].get("content") and result["candidates"][0]["content"].get("parts"):
            return result["candidates"][0]["content"]["parts"][0].get("text", "Não foi possível gerar a mensagem via Gemini.")
        else:
            print(f"DEBUG Gemini Response (estrutura inesperada): {result}")
            return "Erro ao extrair texto da resposta do Gemini (estrutura inesperada)."

    except requests.exceptions.Timeout:
        print("Erro na chamada da API Gemini: Timeout")
        return "Erro ao comunicar com a API Gemini: Timeout."
    except requests.exceptions.RequestException as e:
        print(f"Erro na chamada da API Gemini: {e}")
        return f"Erro ao comunicar com a API Gemini: {e}"
    except Exception as e:
        print(f"Erro inesperado ao processar resposta do Gemini: {e}")
        return "Erro inesperado ao processar mensagem do Gemini."


def _enviar_mensagem_whatsapp_zapi(settings, mensagem):
    zapi_url = f"https://api.z-api.io/instances/{settings['zapi_instancia']}/token/{settings['zapi_token_instancia']}/send-text"
    
    destinatarios = settings.get('whatsapp_para', '').split(',')
    sucessos = 0
    falhas = 0
    mensagens_erro = []

    for destinatario in destinatarios:
        phone_number = destinatario.strip()
        if not phone_number:
            continue
        
        payload = {
            "phone": phone_number,
            "message": mensagem
        }
        headers = {"Content-Type": "application/json"}
        if settings.get('zapi_client_token'): 
            headers['Client-Token'] = settings['zapi_client_token']

        try:
            response = requests.post(zapi_url, headers=headers, json=payload, timeout=30) 
            print(f"Z-API response para {phone_number}: {response.status_code} - {response.text}")
            if response.status_code == 200 or response.status_code == 201: 
                 sucessos += 1
            else:
                falhas += 1
                mensagens_erro.append(f"Falha Z-API para {phone_number}: {response.status_code} - {response.text}")
        except requests.exceptions.Timeout:
            print(f"Erro na chamada da Z-API para {phone_number}: Timeout")
            falhas += 1
            mensagens_erro.append(f"Timeout na Z-API para {phone_number}")
        except requests.exceptions.RequestException as e:
            print(f"Erro na chamada da Z-API para {phone_number}: {e}")
            falhas += 1
            mensagens_erro.append(f"Erro de comunicação com Z-API para {phone_number}: {e}")
        except Exception as e:
            print(f"Erro inesperado ao enviar WhatsApp para {phone_number}: {e}")
            falhas += 1
            mensagens_erro.append(f"Erro inesperado para {phone_number}: {e}")
            
    return sucessos, falhas, mensagens_erro


@app.route('/enviar_notificacao_whatsapp_manual', methods=['POST'])
@login_required
def enviar_notificacao_whatsapp_manual():
    settings = utils.load_notification_settings()
    criterio_selecionado = request.form.get('criterio_wpp_manual', settings['criterio_wpp_manual'])

    if not all([settings.get('zapi_instancia'), settings.get('zapi_token_instancia'), settings.get('whatsapp_para')]):
        return jsonify({"success": False, "message": "Configurações da Z-API (Instância, Token, Destinatários) incompletas."}), 400
    
    if not settings.get('gemini_api_key'):
         return jsonify({"success": False, "message": "API Key do Gemini não configurada."}), 400

    dias_limite = None
    apenas_vencidos = False
    if "30 dias" in criterio_selecionado: dias_limite = 30
    elif "45 dias" in criterio_selecionado: dias_limite = 45
    elif "90 dias" in criterio_selecionado: dias_limite = 90
    elif "apenas equipamentos vencidos" in criterio_selecionado: apenas_vencidos = True
    
    equipamentos_todos = db.fetch_all_equipamentos_completos()
    equipamentos_para_notificar = []

    if equipamentos_todos:
        for equip_row in equipamentos_todos:
            equip = dict(equip_row)
            if not equip.get('ativo') or equip.get('em_calibracao'):
                continue
            
            dias_venc, _, _ = utils.calcular_dias_para_vencimento(
                equip.get('proxima_data_calibracao'),
                equip.get('ativo'),
                equip.get('em_calibracao')
            )
            if dias_venc is not None:
                if apenas_vencidos and dias_venc <= 0:
                    equipamentos_para_notificar.append(equip)
                elif dias_limite is not None and 0 < dias_venc <= dias_limite: 
                    equipamentos_para_notificar.append(equip)
                elif dias_limite is not None and dias_venc <= 0 : 
                    equipamentos_para_notificar.append(equip)

    if not equipamentos_para_notificar:
        return jsonify({"success": True, "message": "Nenhum equipamento encontrado para notificação WhatsApp com o critério selecionado."})

    tabela_texto = _gerar_tabela_texto_para_whatsapp(equipamentos_para_notificar, settings.get('campos_tabela', {}))
    
    if not tabela_texto:
         return jsonify({"success": True, "message": "Nenhum dado de equipamento para gerar a mensagem."})

    mensagem_gerada_gemini = _gerar_mensagem_whatsapp_com_gemini(tabela_texto, settings['gemini_api_key'], settings)
    
    if "Erro" in mensagem_gerada_gemini: 
        return jsonify({"success": False, "message": mensagem_gerada_gemini}), 500

    corpo_template_whatsapp = settings.get('corpo_template_whatsapp', "{tabela_equipamentos_texto}")
    mensagem_final_whatsapp = corpo_template_whatsapp.replace("{tabela_equipamentos_texto}", mensagem_gerada_gemini)

    sucessos, falhas, erros_envio = _enviar_mensagem_whatsapp_zapi(settings, mensagem_final_whatsapp)

    if sucessos > 0 and falhas == 0:
        return jsonify({"success": True, "message": f"{sucessos} mensagem(ns) de WhatsApp enviada(s) com sucesso!"})
    elif sucessos > 0 and falhas > 0:
        return jsonify({"success": True, "message": f"{sucessos} mensagem(ns) enviada(s), {falhas} falha(s). Detalhes: {'; '.join(erros_envio)}"})
    else:
        return jsonify({"success": False, "message": f"Nenhuma mensagem de WhatsApp enviada. Falhas: {falhas}. Detalhes: {'; '.join(erros_envio)}"}), 500


# --- Rotas de Exportação Excel ---
@app.route('/exportar_geral_excel')
@login_required
def exportar_geral_excel():
    search_query = request.args.get('search', '') 
    if search_query:
        equipamentos_data_raw = db.search_equipamentos(search_query)
    else:
        equipamentos_data_raw = db.fetch_all_equipamentos_completos()

    if not equipamentos_data_raw:
        flash("Nenhum equipamento para exportar.", "info")
        return redirect(url_for('lista_equipamentos'))

    output = BytesIO()
    wb = openpyxl.Workbook() 
    
    ws_equip = wb.active
    ws_equip.title = "Equipamentos"
    headers_equip = ["ID", "Nome", "TAG", "Fabricante", "Modelo", "Nº Série", "Status", "Localização", "Obs. Equip.", 
                     "Tipo Equipamento", "Faixa de Uso",
                     "Últ. Nº Cert.", "Últ. Data Cal.", "Próx. Data Cal.",
                     "Últ. Res. Geral Certificado",
                     "Ativo", "Requer Calibração", "Em Calibração", "Destino Inativo"]
    ws_equip.append(headers_equip)
    for col_num, header_title in enumerate(headers_equip, 1):
        cell = ws_equip.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for equip_row in equipamentos_data_raw:
        equip = dict(equip_row)
        dados_linha = [
            equip.get('id'), equip.get('nome'), equip.get('tag'), equip.get('fabricante'), equip.get('modelo'), equip.get('numero_serie'), 
            equip.get('status'), equip.get('localizacao'), equip.get('observacoes_equipamento'),
            equip.get('tipo_equipamento_nome'), equip.get('faixa_de_uso'),
            equip.get('ultimo_numero_certificado'), 
            utils.format_date_for_display(equip.get('ultima_data_calibracao')), 
            utils.format_date_for_display(equip.get('proxima_data_calibracao')),
            equip.get('ultimo_resultado_geral_certificado'),
            "Sim" if equip.get('ativo') else "Não",
            "Sim" if equip.get('requer_calibracao') else "Não",
            "Sim" if equip.get('em_calibracao') else "Não",
            equip.get('destino_inativo') or ""
        ]
        ws_equip.append(dados_linha)
    
    ws_analises = wb.create_sheet(title="Historico Análises")
    headers_analises = ["ID Equip.", "Nome Equip.", "ID Análise", "Data Reg. Sistema", "Data Análise", "Responsável",
                        "Nº Cert.", "Data Cal. (Cert.)", "Próx. Cal. (Cert.)",
                        "Resultado Geral Cert.", "Obs. Análise"]
    ws_analises.append(headers_analises)
    for col_num, header_title in enumerate(headers_analises, 1):
        ws_analises.cell(row=1, column=col_num).font = Font(bold=True)
        ws_analises.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    ws_pontos = wb.create_sheet(title="Pontos Analisados")
    headers_pontos = ["ID Análise", "Nº Cert. Análise", "Nome Ponto", "Símbolo", "Valor Nominal",
                      "Amplitude A", "Desvio B", "Regra Aplicada", "Resultado Ponto", "Obs. Ponto"]
    ws_pontos.append(headers_pontos)
    for col_num, header_title in enumerate(headers_pontos, 1):
        ws_pontos.cell(row=1, column=col_num).font = Font(bold=True)
        ws_pontos.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    ws_anexos = wb.create_sheet(title="Anexos")
    headers_anexos = ["ID Análise", "Nº Cert. Análise", "Nome Original Anexo", "Caminho Armazenado", "Data Anexo"]
    ws_anexos.append(headers_anexos)
    for col_num, header_title in enumerate(headers_anexos, 1):
        ws_anexos.cell(row=1, column=col_num).font = Font(bold=True)
        ws_anexos.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    if equipamentos_data_raw:
        for equip_row in equipamentos_data_raw:
            equip = dict(equip_row)
            equip_id = equip.get('id')
            equip_nome = equip.get('nome')
            
            hist_analises = db.fetch_analises_by_equipamento_id(equip_id, app_utils_instance=utils) 
            for analise_dict_from_db in hist_analises:
                analise = dict(analise_dict_from_db) 
                ws_analises.append([
                    equip_id, equip_nome, analise.get('id'),
                    analise.get('data_registro_sistema_fmt'), 
                    analise.get('data_analise_manual_fmt'),
                    analise.get('responsavel_analise'),
                    analise.get('numero_certificado_analisado'),
                    analise.get('data_calibracao_analisada_fmt'),
                    analise.get('data_prox_calibracao_analisada_fmt'),
                    analise.get('resultado_geral_certificado'),
                    analise.get('observacoes_analise')
                ])

                pontos_da_analise = db.fetch_pontos_by_analise_id(analise['id'])
                for ponto_row in pontos_da_analise:
                    ponto = dict(ponto_row)
                    ws_pontos.append([
                        analise.get('id'), analise.get('numero_certificado_analisado'),
                        ponto.get('nome_ponto'), ponto.get('simbolo_ponto'), ponto.get('valor_nominal_ponto'),
                        ponto.get('amplitude_A_ponto'), ponto.get('desvio_B_ponto'),
                        ponto.get('regra_aplicada_ponto'), ponto.get('resultado_ponto'),
                        ponto.get('observacoes_ponto')
                    ])
                
                anexos_da_an = db.fetch_anexos_by_analise_id(analise['id'])
                for anexo_row in anexos_da_an:
                    anexo = dict(anexo_row)
                    ws_anexos.append([
                        analise.get('id'), 
                        analise.get('numero_certificado_analisado'),
                        anexo.get('nome_arquivo_original'),
                        anexo.get('caminho_relativo_armazenado'),
                        utils.format_date_for_display(anexo.get('data_anexo'))
                ])
            for cell in column_cells_tuple:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                except:
                    pass
            adjusted_width = (max_length + 2 if max_length > 0 else 12) 
            ws.column_dimensions[column_letter].width = adjusted_width
            
    wb.save(output) 
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=relatorio_equipamentos_completo.xlsx"}
    )


@app.route('/exportar_individual_excel/<int:equip_id>')
@login_required
def exportar_individual_excel(equip_id):
    equip_data = db.fetch_equipamento_completo_by_id(equip_id)
    if not equip_data:
        flash("Equipamento não encontrado.", "danger")
        return redirect(url_for('lista_equipamentos'))

    equip = dict(equip_data)
    
    output = BytesIO()
    wb = openpyxl.Workbook() 
    
    ws_equip = wb.active
    ws_equip.title = "Equipamentos"
    headers_equip = ["ID", "Nome", "TAG", "Fabricante", "Modelo", "Nº Série", "Status", "Localização", "Obs. Equip.", 
                     "Tipo Equipamento", "Faixa de Uso",
                     "Últ. Nº Cert.", "Últ. Data Cal.", "Próx. Data Cal.",
                     "Últ. Res. Geral Certificado",
                     "Ativo", "Requer Calibração", "Em Calibração", "Destino Inativo"]
    ws_equip.append(headers_equip)
    for col_num, header_title in enumerate(headers_equip, 1):
        cell = ws_equip.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    dados_linha_equip = [
        equip.get('id'), equip.get('nome'), equip.get('tag'), equip.get('fabricante'), equip.get('modelo'), equip.get('numero_serie'), 
        equip.get('status'), equip.get('localizacao'), equip.get('observacoes_equipamento'),
        equip.get('tipo_equipamento_nome'), equip.get('faixa_de_uso'),
        equip.get('ultimo_numero_certificado'), 
        utils.format_date_for_display(equip.get('ultima_data_calibracao')), 
        utils.format_date_for_display(equip.get('proxima_data_calibracao')),
        equip.get('ultimo_resultado_geral_certificado'),
        "Sim" if equip.get('ativo') else "Não",
        "Sim" if equip.get('requer_calibracao') else "Não",
        "Sim" if equip.get('em_calibracao') else "Não",
        equip.get('destino_inativo') or ""
    ]
    ws_equip.append(dados_linha_equip)
    
    ws_analises = wb.create_sheet(title="Historico Análises")
    headers_analises = ["ID Equip.", "Nome Equip.", "ID Análise", "Data Reg. Sistema", "Data Análise", "Responsável",
                        "Nº Cert.", "Data Cal. (Cert.)", "Próx. Cal. (Cert.)",
                        "Resultado Geral Cert.", "Obs. Análise"]
    ws_analises.append(headers_analises)
    for col_num, header_title in enumerate(headers_analises, 1):
        ws_analises.cell(row=1, column=col_num).font = Font(bold=True)
        ws_analises.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    ws_pontos = wb.create_sheet(title="Pontos Analisados")
    headers_pontos = ["ID Análise", "Nº Cert. Análise", "Nome Ponto", "Símbolo", "Valor Nominal",
                      "Amplitude A", "Desvio B", "Regra Aplicada", "Resultado Ponto", "Obs. Ponto"]
    ws_pontos.append(headers_pontos)
    for col_num, header_title in enumerate(headers_pontos, 1):
        ws_pontos.cell(row=1, column=col_num).font = Font(bold=True)
       
        ws_pontos.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    ws_anexos = wb.create_sheet(title="Anexos")
    headers_anexos = ["ID Análise", "Nº Cert. Análise", "Nome Original Anexo", "Caminho Armazenado", "Data Anexo"]
    ws_anexos.append(headers_anexos)
    for col_num, header_title in enumerate(headers_anexos, 1):
        ws_anexos.cell(row=1, column=col_num).font = Font(bold=True)
        ws_anexos.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")
    
    equip_id_atual = equip.get('id')
    equip_nome_atual = equip.get('nome')
    
    hist_analises_individuais = db.fetch_analises_by_equipamento_id(equip_id_atual, app_utils_instance=utils) 
    for analise_dict_from_db in hist_analises_individuais:
        analise = dict(analise_dict_from_db)
        ws_analises.append([
            equip_id_atual, equip_nome_atual, analise.get('id'),
            analise.get('data_registro_sistema_fmt'), 
            analise.get('data_analise_manual_fmt'),
            analise.get('responsavel_analise'),
            analise.get('numero_certificado_analisado'),
            analise.get('data_calibracao_analisada_fmt'),
            analise.get('data_prox_calibracao_analisada_fmt'),
            analise.get('resultado_geral_certificado'),
            analise.get('observacoes_analise')
        ])

        pontos_da_analise = db.fetch_pontos_by_analise_id(analise['id'])
        for ponto_row in pontos_da_analise:
            ponto = dict(ponto_row)
            ws_pontos.append([
                analise.get('id'), analise.get('numero_certificado_analisado'),
                ponto.get('nome_ponto'), ponto.get('simbolo_ponto'), ponto.get('valor_nominal_ponto'),
                ponto.get('amplitude_A_ponto'), ponto.get('desvio_B_ponto'),
                ponto.get('regra_aplicada_ponto'), ponto.get('resultado_ponto'),
                ponto.get('observacoes_ponto')
            ])

        anexos_da_an = db.fetch_anexos_by_analise_id(analise['id'])
        for anexo_row in anexos_da_an:
            anexo = dict(anexo_row)
            ws_anexos.append([
                analise.get('id'), analise.get('numero_certificado_analisado'),
                anexo.get('nome_arquivo_original'),
                anexo.get('caminho_relativo_armazenado'),
                utils.format_date_for_display(anexo.get('data_anexo')) 
            ])

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_idx, column_cells_tuple in enumerate(ws.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            for cell in column_cells_tuple:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2 if max_length > 0 else 12) 
            ws.column_dimensions[column_letter].width = adjusted_width
            
    wb.save(output) 
    output.seek(0)
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=relatorio_equip_{equip_id}_{equip.get('nome', '')}.xlsx"}
 )
@app.route('/equipamento/<int:equip_id>/analise/nova_form')
def nova_analise_form(equip_id):
    equip = db.fetch_equipamento_completo_by_id(equip_id)
    if not equip:
        return jsonify({"error": "Equipamento não encontrado"}), 404

    # Adicionei a chamada get_db() aqui também, caso seja usada nesta rota.
    if 'db' not in g:
        g.db = DatabaseManager(DB_FULL_PATH)
    db = g.db

    return render_template('nova_analise.html', equipamento=equip)

if __name__ == "__main__":
    app.run(debug=True)